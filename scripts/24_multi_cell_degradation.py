"""Multi-cell degradation analysis across all available experimental data.

Cells:
- NEWAREA: 434 cycles, 155 mAh, 62.8% fade (heavy degradation)
- NEWAREB: 325 cycles, 153 mAh, 15.8% fade (moderate degradation)
- TVC: 221 cycles, 142 mAh, ~2% fade (minimal degradation, 45°C)

Key analysis:
1. Capacity fade curves and degradation phases
2. ICA evolution and peak tracking
3. Degradation mode classification
4. Early-life prediction benchmark
"""

import os
os.environ["MKL_THREADING_LAYER"] = "GNU"

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from pathlib import Path
from scipy.ndimage import gaussian_filter1d
from sklearn.linear_model import Ridge
from sklearn.preprocessing import StandardScaler
import openpyxl
import logging

from src.data.loader import ExperimentalDataLoader

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
logger = logging.getLogger(__name__)

OUT = Path("outputs/degradation")
OUT.mkdir(parents=True, exist_ok=True)


def load_newarea():
    loader = ExperimentalDataLoader()
    data = loader.load_neware_xlsx("/root/data/raw/exapmles/NEWAREA_1205XXL01006.xlsx")
    return extract_features(data, "NEWAREA")


def load_newareb():
    wb = openpyxl.load_workbook("/root/data/raw/exapmles/NEWAREB_1122XXL01002.xlsx", data_only=True)
    ws = wb["Cycle_7_1_2"]
    rows = list(ws.iter_rows(values_only=True))

    cycles, caps = [], []
    for r in rows[1:]:
        if r[1] is not None and r[3] is not None:
            try:
                cyc = int(float(r[1]))
                cap = float(r[3])
                if cap > 0.001:
                    cycles.append(cyc)
                    caps.append(cap)
            except:
                pass
    wb.close()

    return np.array(cycles), np.array(caps), "NEWAREB"


def load_tvc():
    wb = openpyxl.load_workbook(
        "/root/data/raw/exapmles/TVC_T3-20251204-1682-Cycle life_1830857_45℃_0.1C_1C.xlsx",
        data_only=True,
    )
    ws = wb["CD_Capacity_Data"]
    rows = list(ws.iter_rows(values_only=True))

    cycles, caps1, caps2, dcir1, dcir2 = [], [], [], [], []

    for r in rows[2:]:
        if r[0] is None:
            continue
        try:
            cycles.append(int(float(r[0])))
            caps1.append(float(r[2]) if r[2] else 0)
            caps2.append(float(r[4]) if r[4] else 0)
        except:
            pass

    ws_dcir = wb["30s Cycle_DCIR"]
    dcir_rows = list(ws_dcir.iter_rows(values_only=True))
    dcir_cycles, dcir1, dcir2 = [], [], []
    for r in dcir_rows[2:]:
        if r[0] is None:
            continue
        try:
            dcir_cycles.append(int(float(r[0])))
            dcir1.append(float(r[1]) if r[1] else 0)
            dcir2.append(float(r[2]) if r[2] else 0)
        except:
            pass
    wb.close()

    return (
        np.array(cycles), np.array(caps1), np.array(caps2),
        np.array(dcir_cycles), np.array(dcir1), np.array(dcir2),
    )


def extract_features(data, name):
    cycles_arr = data["cycle"]
    current = data["current"]
    voltage = data["voltage"]
    time_arr = data["time"]

    unique = np.unique(cycles_arr)
    result_cycles, result_caps, v_means, v_ranges, durations, ica_peaks = [], [], [], [], [], []

    for c in unique:
        mask = (cycles_arr == c) & (current < -0.005)
        idx = np.where(mask)[0]
        if len(idx) < 15:
            continue
        v_seg = voltage[idx]
        i_seg = current[idx]
        t_seg = time_arr[idx]
        valid = ~np.isnan(v_seg)
        v_seg, i_seg, t_seg = v_seg[valid], i_seg[valid], t_seg[valid]
        if len(v_seg) < 15:
            continue

        dt = np.diff(t_seg)
        if dt.sum() < 10:
            continue
        cap = -np.sum(i_seg[:-1] * dt) / 3600
        if cap < 0.001:
            continue

        result_cycles.append(int(c))
        result_caps.append(cap)
        v_means.append(v_seg.mean())
        v_ranges.append(v_seg.max() - v_seg.min())
        durations.append(t_seg[-1] - t_seg[0])

        Q = np.cumsum(np.concatenate([[0], -i_seg[:-1] * dt / 3600]))
        Q_s = gaussian_filter1d(Q, sigma=2)
        V_s = gaussian_filter1d(v_seg, sigma=2)
        dV = np.diff(V_s)
        dQ = np.diff(Q_s)
        ok = np.abs(dV) > 1e-5
        if ok.sum() > 10:
            dQdV = gaussian_filter1d(dQ[ok] / dV[ok], sigma=3)
            ica_peaks.append(float(np.max(dQdV)))
        else:
            ica_peaks.append(0.0)

    return {
        "cycles": np.array(result_cycles),
        "caps": np.array(result_caps),
        "v_means": np.array(v_means),
        "v_ranges": np.array(v_ranges),
        "durations": np.array(durations),
        "ica_peaks": np.array(ica_peaks),
        "name": name,
    }


def plot_multi_cell_comparison(cells, tvc_data):
    fig, axes = plt.subplots(2, 3, figsize=(18, 11))

    colors = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728"]

    # 1. Capacity retention
    ax = axes[0, 0]
    for i, cell in enumerate(cells):
        ax.plot(cell["cycles"], cell["caps"] / cell["caps"].max() * 100, color=colors[i], label=cell["name"], linewidth=1.5)
    if tvc_data is not None:
        cycles_t, c1, c2 = tvc_data[0], tvc_data[1], tvc_data[2]
        ax.plot(cycles_t, c1 / c1.max() * 100, color=colors[2], label="TVC Cell1", linewidth=1.5)
        ax.plot(cycles_t, c2 / c2.max() * 100, color=colors[3], label="TVC Cell2", linewidth=1.5)
    ax.axhline(80, color="red", linestyle="--", alpha=0.4, label="80% EOL")
    ax.set_xlabel("Cycle"); ax.set_ylabel("Capacity retention (%)")
    ax.set_title("Capacity Fade Comparison"); ax.legend(fontsize=8); ax.grid(True, alpha=0.3)

    # 2. Fade rate comparison (per-cycle)
    ax = axes[0, 1]
    for i, cell in enumerate(cells):
        fade = (cell["caps"][0] - cell["caps"]) / cell["caps"][0] * 100
        ax.plot(cell["cycles"], fade, color=colors[i], label=cell["name"], linewidth=1.5)
    ax.set_xlabel("Cycle"); ax.set_ylabel("Cumulative fade (%)")
    ax.set_title("Cumulative Capacity Fade"); ax.legend(fontsize=8); ax.grid(True, alpha=0.3)

    # 3. Mean voltage
    ax = axes[0, 2]
    for i, cell in enumerate(cells):
        ax.plot(cell["cycles"], cell["v_means"], color=colors[i], label=cell["name"], linewidth=1.5)
    ax.set_xlabel("Cycle"); ax.set_ylabel("Mean discharge voltage (V)")
    ax.set_title("Voltage Evolution"); ax.legend(fontsize=8); ax.grid(True, alpha=0.3)

    # 4. Voltage range
    ax = axes[1, 0]
    for i, cell in enumerate(cells):
        ax.plot(cell["cycles"], cell["v_ranges"] * 1000, color=colors[i], label=cell["name"], linewidth=1.5)
    ax.set_xlabel("Cycle"); ax.set_ylabel("Voltage range (mV)")
    ax.set_title("Voltage Range Evolution"); ax.legend(fontsize=8); ax.grid(True, alpha=0.3)

    # 5. ICA peak height
    ax = axes[1, 1]
    for i, cell in enumerate(cells):
        ax.plot(cell["cycles"], cell["ica_peaks"], color=colors[i], label=cell["name"], linewidth=1.5)
    ax.set_xlabel("Cycle"); ax.set_ylabel("ICA peak height (dQ/dV)")
    ax.set_title("ICA Peak Evolution"); ax.legend(fontsize=8); ax.grid(True, alpha=0.3)

    # 6. DCIR (TVC only)
    ax = axes[1, 2]
    if tvc_data is not None and len(tvc_data) >= 6:
        dcir_c, d1, d2 = tvc_data[3], tvc_data[4], tvc_data[5]
        ax.plot(dcir_c, d1, color=colors[2], label="TVC Cell1", linewidth=1.5)
        ax.plot(dcir_c, d2, color=colors[3], label="TVC Cell2", linewidth=1.5)
        ax.set_xlabel("Cycle"); ax.set_ylabel("DCIR (mΩ)")
        ax.set_title("DC Resistance Evolution (TVC)")
        ax.legend(fontsize=8); ax.grid(True, alpha=0.3)

    fig.suptitle("Multi-Cell Degradation Comparison", fontsize=14)
    fig.tight_layout()
    fig.savefig(OUT / "multi_cell_comparison.png", dpi=150)
    plt.close(fig)


def plot_degradation_phases(cell):
    name = cell["name"]
    caps = cell["caps"]
    cycles = cell["cycles"]
    c_max = caps.max()

    fig, axes = plt.subplots(1, 3, figsize=(16, 5))

    # Capacity with phases
    ax = axes[0]
    ax.plot(cycles, caps * 1000, "b-", linewidth=1)
    ax.axhline(c_max * 0.8 * 1000, color="r", linestyle="--", alpha=0.5, label="80% SOH")

    # Linear fit
    n = len(caps)
    if n > 50:
        early = min(300, n * 2 // 3)
        coef = np.polyfit(cycles[:early], caps[:early] * 1000, 1)
        ax.plot(cycles, np.polyval(coef, cycles), "g--", alpha=0.5, label=f"Linear fit (1-{early})")
    ax.set_xlabel("Cycle"); ax.set_ylabel("Capacity (mAh)")
    ax.set_title(f"{name}: Capacity Fade"); ax.legend(fontsize=8); ax.grid(True, alpha=0.3)

    # dQ/dC (capacity fade rate)
    ax = axes[1]
    dc = np.gradient(caps * 1000, cycles)
    dc_smooth = gaussian_filter1d(dc, sigma=5)
    ax.plot(cycles, dc_smooth, "r-")
    ax.axhline(0, color="k", alpha=0.3)
    ax.set_xlabel("Cycle"); ax.set_ylabel("dQ/dC (mAh/cycle)")
    ax.set_title(f"{name}: Fade Rate"); ax.grid(True, alpha=0.3)

    # d²Q/dC² (acceleration)
    ax = axes[2]
    d2c = np.gradient(dc_smooth, cycles)
    d2c_smooth = gaussian_filter1d(d2c, sigma=5)
    ax.plot(cycles, d2c_smooth, "m-")
    ax.axhline(0, color="k", alpha=0.3)
    ax.set_xlabel("Cycle"); ax.set_ylabel("d²Q/dC² (mAh/cycle²)")
    ax.set_title(f"{name}: Fade Acceleration"); ax.grid(True, alpha=0.3)

    fig.tight_layout()
    fig.savefig(OUT / f"{name.lower()}_phases.png", dpi=150)
    plt.close(fig)


def early_life_prediction_benchmark(cells):
    fig, axes = plt.subplots(1, len(cells), figsize=(6 * len(cells), 5))
    if len(cells) == 1:
        axes = [axes]

    for ci, cell in enumerate(cells):
        name = cell["name"]
        caps = cell["caps"]
        cycles = cell["cycles"]

        feat_keys = ["caps", "v_means", "v_ranges", "durations", "ica_peaks"]
        X = np.column_stack([cell[k] for k in feat_keys])
        y = caps

        results = []
        for early_pct in [0.05, 0.10, 0.15, 0.20, 0.30]:
            n_train = max(3, int(len(caps) * early_pct))
            scaler = StandardScaler()
            X_train = scaler.fit_transform(X[:n_train])
            X_all = scaler.transform(X)
            model = Ridge(alpha=1.0)
            model.fit(X_train, y[:n_train])
            y_pred = model.predict(X_all)
            rmse = np.sqrt(np.mean((y_pred[n_train:] - y[n_train:]) ** 2)) * 1000
            results.append((early_pct, n_train, rmse, y_pred))

        ax = axes[ci]
        ax.plot(cycles, caps * 1000, "b-", linewidth=2, label="Actual")
        colors_pred = ["#ff7f0e", "#2ca02c", "#d62728", "#9467bd", "#8c564b"]
        for i, (pct, n_train, rmse, y_pred) in enumerate(results):
            ax.plot(cycles, y_pred * 1000, "--", color=colors_pred[i], alpha=0.6,
                   label=f"{pct*100:.0f}% ({n_train} cyc), RMSE={rmse:.1f}mAh")
        ax.set_xlabel("Cycle"); ax.set_ylabel("Capacity (mAh)")
        ax.set_title(f"{name}"); ax.legend(fontsize=7); ax.grid(True, alpha=0.3)

    fig.suptitle("Early-Life Capacity Prediction", fontsize=14)
    fig.tight_layout()
    fig.savefig(OUT / "early_life_prediction_benchmark.png", dpi=150)
    plt.close(fig)

    logger.info("Early-life prediction benchmark:")
    for cell in cells:
        caps = cell["caps"]
        cycles = cell["cycles"]
        X = np.column_stack([cell[k] for k in ["caps", "v_means", "v_ranges", "durations", "ica_peaks"]])
        y = caps
        for pct in [0.10, 0.20]:
            n_train = max(3, int(len(caps) * pct))
            scaler = StandardScaler()
            model = Ridge(alpha=1.0)
            model.fit(scaler.fit_transform(X[:n_train]), y[:n_train])
            y_pred = model.predict(scaler.transform(X))
            rmse = np.sqrt(np.mean((y_pred[n_train:] - y[n_train:]) ** 2)) * 1000
            logger.info(f"  {cell['name']}: {pct*100:.0f}% train ({n_train} cyc) → RMSE={rmse:.1f} mAh ({rmse/caps[0]/10:.2f}%)")


def main():
    logger.info("Loading all experimental data...")

    # Load NEWAREA (full feature extraction)
    cell_a = load_newarea()
    logger.info(f"NEWAREA: {len(cell_a['cycles'])} cycles, {cell_a['caps'].max()*1000:.1f} mAh → {cell_a['caps'].min()*1000:.1f} mAh ({cell_a['caps'].min()/cell_a['caps'].max()*100:.1f}%)")

    # Load NEWAREB (summary only)
    cycles_b, caps_b, name_b = load_newareb()
    cell_b = {"cycles": cycles_b, "caps": caps_b, "v_means": np.zeros_like(caps_b),
              "v_ranges": np.zeros_like(caps_b), "durations": np.zeros_like(caps_b),
              "ica_peaks": np.zeros_like(caps_b), "name": name_b}
    logger.info(f"NEWAREB: {len(cycles_b)} cycles, {caps_b.max()*1000:.1f} mAh → {caps_b.min()*1000:.1f} mAh ({caps_b.min()/caps_b.max()*100:.1f}%)")

    # Load TVC (summary + DCIR)
    tvc_data = load_tvc()
    logger.info(f"TVC: {len(tvc_data[0])} cycles, Cell1 {tvc_data[1].max()*1000:.1f} → {tvc_data[1][-1]*1000:.1f} mAh")

    # Multi-cell comparison
    logger.info("Plotting multi-cell comparison...")
    plot_multi_cell_comparison([cell_a, cell_b], tvc_data)

    # Degradation phases for each cell with full data
    for cell in [cell_a]:
        plot_degradation_phases(cell)

    # Early-life prediction
    logger.info("Running early-life prediction benchmark...")
    cells_with_features = [cell for cell in [cell_a] if cell["v_means"].sum() > 0]
    if cells_with_features:
        early_life_prediction_benchmark(cells_with_features)

    # Degradation mode classification
    logger.info("\n" + "=" * 60)
    logger.info("DEGRADATION MODE CLASSIFICATION")
    logger.info("=" * 60)

    # NEWAREA analysis
    caps_a = cell_a["caps"]
    cycles_a = cell_a["cycles"]
    c_max = caps_a.max()

    # Linear phase
    early_mask = cycles_a <= 300
    if early_mask.sum() > 10:
        rate = np.polyfit(cycles_a[early_mask], caps_a[early_mask] * 1000, 1)
        logger.info(f"\nNEWAREA Phase Analysis:")
        logger.info(f"  Linear phase (1-300): {abs(rate[0]):.2f} mAh/cycle ({abs(rate[0])/c_max/10:.4f}%/cycle)")

    late_mask = cycles_a >= 350
    if late_mask.sum() > 5:
        rate2 = np.polyfit(cycles_a[late_mask], caps_a[late_mask] * 1000, 1)
        logger.info(f"  Rapid phase (350-434): {abs(rate2[0]):.2f} mAh/cycle ({abs(rate2[0])/c_max/10:.4f}%/cycle)")
        if rate[0] != 0:
            logger.info(f"  Acceleration: {abs(rate2[0]/rate[0]):.1f}x")

    eol_idx = np.where(caps_a < c_max * 0.8)[0]
    if len(eol_idx) > 0:
        logger.info(f"  80% SOH at cycle {int(cycles_a[eol_idx[0]])}")

    logger.info(f"\n  Pattern: Linear fade → Knee → Accelerating failure")
    logger.info(f"  Inferred mechanism: SEI growth (linear) → Li plating onset → rapid LAM/LLI")

    # NEWAREB
    logger.info(f"\nNEWAREB: {caps_b[-1]/caps_b.max()*100:.1f}% retention over {len(cycles_b)} cycles")
    logger.info(f"  Pattern: Slow linear fade ({(1-caps_b[-1]/caps_b.max())*100/len(cycles_b):.4f}%/cycle)")
    logger.info(f"  Inferred mechanism: Predominantly SEI growth")

    # TVC
    dcir_growth = tvc_data[4][-1] / tvc_data[4][0]
    logger.info(f"\nTVC (45°C): {tvc_data[1][-1]/tvc_data[1][0]*100:.1f}% retention over {len(tvc_data[0])} cycles")
    logger.info(f"  DCIR growth: {dcir_growth:.2f}x ({tvc_data[4][0]:.0f} → {tvc_data[4][-1]:.0f} mΩ)")
    logger.info(f"  Pattern: Very slow fade with measurable resistance growth at elevated temperature")
    logger.info(f"  Inferred mechanism: Temperature-accelerated SEI growth")

    logger.info(f"\nAll outputs saved to {OUT}/")
    logger.info("Multi-cell degradation analysis complete!")


if __name__ == "__main__":
    main()
